from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, make_response
from app.models import Project, Estimate, Assembly, AssemblyPart, Parts, PartCategory, EstimateComponent, Motor
from app import db
from datetime import datetime
from sqlalchemy.orm import joinedload
from sqlalchemy import func, select
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from decimal import Decimal
import re
from collections import defaultdict
import uuid

bp = Blueprint('projects', __name__)

@bp.route('/')
def list_projects():
    """List all projects with pagination and optimized loading"""
    # Get pagination parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)  # Default 50 projects per page

    # Get filter parameters
    status_filter = request.args.get('status', '')
    client_filter = request.args.get('client', '')

    # Create subquery to count estimates per project
    estimate_count_subquery = (
        select(func.count(Estimate.estimate_id))
        .where(Estimate.project_id == Project.project_id)
        .correlate(Project)
        .scalar_subquery()
    )

    # Build query - select Project with estimate count
    query = db.session.query(
        Project,
        estimate_count_subquery.label('estimate_count')
    )

    # Apply filters
    if status_filter:
        query = query.filter(Project.status == status_filter)
    if client_filter:
        query = query.filter(Project.client_name.ilike(f'%{client_filter}%'))

    # Order by most recently updated
    query = query.order_by(Project.updated_at.desc())

    # Paginate
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    # Extract projects and their counts
    projects_with_counts = []
    for item in pagination.items:
        project, estimate_count = item
        project.estimate_count = estimate_count  # Attach count to project object
        projects_with_counts.append(project)

    # Hardcode common statuses instead of querying database
    statuses = ['Draft', 'Sent', 'Won', 'Lost', 'On Hold']

    return render_template('projects/list.html',
                          projects=projects_with_counts,
                          pagination=pagination,
                          statuses=statuses,
                          current_status=status_filter,
                          current_client=client_filter)

@bp.route('/create', methods=['GET', 'POST'])
def create_project():
    """Create a new project"""
    if request.method == 'POST':
        try:
            project = Project(
                project_name=request.form['project_name'],
                client_name=request.form['client_name'],
                description=request.form.get('description', ''),
                status=request.form.get('status', 'Draft'),
                revision=request.form.get('revision', ''),
                remarks=request.form.get('remarks', '')
            )
            
            db.session.add(project)
            db.session.commit()
            
            flash(f'Project "{project.project_name}" created successfully!', 'success')
            return redirect(url_for('projects.detail_project', project_id=project.project_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating project: {str(e)}', 'error')
    
    return render_template('projects/create.html')

@bp.route('/<int:project_id>')
def detail_project(project_id):
    """Show project details with all estimates"""
    project = Project.query.get_or_404(project_id)

    # Separate standard and optional estimates
    standard_estimates = Estimate.query.filter_by(project_id=project_id, is_optional=False).order_by(Estimate.sort_order, Estimate.created_at.desc()).all()
    optional_estimates = Estimate.query.filter_by(project_id=project_id, is_optional=True).order_by(Estimate.sort_order, Estimate.created_at.desc()).all()

    # Calculate category costs for chart (standard estimates only)
    category_costs = defaultdict(float)

    # Labor rates
    ENGINEERING_RATE = 145.0
    PANEL_SHOP_RATE = 125.0

    for estimate in standard_estimates:
        # Process assemblies
        for assembly in estimate.assemblies:
            for assembly_part in assembly.assembly_parts:
                part = assembly_part.part
                if part:
                    # Get category
                    category = 'Uncategorized'
                    if hasattr(part, 'category_id') and part.category_id:
                        try:
                            cat = PartCategory.query.get(part.category_id)
                            if cat:
                                category = cat.name
                        except:
                            pass

                    # Calculate cost
                    qty = float(assembly_part.quantity) if assembly_part.quantity else 0
                    unit_cost = float(part.current_price) if part.current_price else 0
                    category_costs[category] += qty * unit_cost

        # Process individual components
        for component in estimate.individual_components:
            category = 'Uncategorized'
            if component.part_id and component.part:
                if hasattr(component.part, 'category_id') and component.part.category_id:
                    try:
                        cat = PartCategory.query.get(component.part.category_id)
                        if cat:
                            category = cat.name
                    except:
                        pass
            elif component.category:
                category = component.category

            qty = float(component.quantity) if component.quantity else 0
            unit_cost = float(component.unit_price) if component.unit_price else 0
            category_costs[category] += qty * unit_cost

    # Add Engineering Hours as a category
    total_engineering_hours = sum(float(est.total_engineering_hours or 0) for est in standard_estimates)
    if total_engineering_hours > 0:
        category_costs['Engineering'] = total_engineering_hours * ENGINEERING_RATE

    # Add Panel Shop Hours as a category
    total_panel_shop_hours = sum(float(est.total_panel_shop_hours or 0) for est in standard_estimates)
    if total_panel_shop_hours > 0:
        category_costs['Panel Shop'] = total_panel_shop_hours * PANEL_SHOP_RATE

    # Get top 15 categories by cost
    top_categories = sorted(category_costs.items(), key=lambda x: x[1], reverse=True)[:15]

    # Calculate total for percentage calculations (materials + engineering + panel shop)
    total_chart_cost = sum(cost for _, cost in category_costs.items())

    return render_template('projects/detail.html',
                         project=project,
                         estimates=standard_estimates,
                         optional_estimates=optional_estimates,
                         top_categories=top_categories,
                         total_chart_cost=total_chart_cost)

@bp.route('/<int:project_id>/edit', methods=['GET', 'POST'])
def edit_project(project_id):
    """Edit an existing project"""
    project = Project.query.get_or_404(project_id)
    
    if request.method == 'POST':
        try:
            project.project_name = request.form['project_name']
            project.client_name = request.form['client_name']
            project.description = request.form.get('description', '')
            project.status = request.form.get('status', 'Draft')
            project.revision = request.form.get('revision', '')
            project.remarks = request.form.get('remarks', '')
            project.is_active = 'is_active' in request.form
            project.updated_at = datetime.utcnow()

            db.session.commit()
            flash('Project updated successfully!', 'success')
            return redirect(url_for('projects.detail_project', project_id=project_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating project: {str(e)}', 'error')
    
    return render_template('projects/edit.html', project=project)

@bp.route('/<int:project_id>/delete', methods=['POST'])
def delete_project(project_id):
    """Delete a project and all associated data (estimates, assemblies, motors, etc.)"""
    project = Project.query.get_or_404(project_id)

    try:
        project_name = project.project_name

        # Count related items before deletion for detailed feedback
        estimates = Estimate.query.filter_by(project_id=project_id).all()
        motors = Motor.query.filter_by(project_id=project_id).all()

        estimate_count = len(estimates)
        motor_count = len(motors)

        # Count assemblies and components across all estimates
        estimate_ids = [e.estimate_id for e in estimates]
        assembly_count = 0
        assembly_part_count = 0
        component_count = 0

        if estimate_ids:
            assemblies = Assembly.query.filter(Assembly.estimate_id.in_(estimate_ids)).all()
            assembly_count = len(assemblies)

            assembly_ids = [a.assembly_id for a in assemblies]
            if assembly_ids:
                assembly_part_count = AssemblyPart.query.filter(AssemblyPart.assembly_id.in_(assembly_ids)).count()

            component_count = EstimateComponent.query.filter(EstimateComponent.estimate_id.in_(estimate_ids)).count()

        # Delete the project (cascade will handle all related data)
        db.session.delete(project)
        db.session.commit()

        # Verify deletion was successful
        remaining_estimates = Estimate.query.filter_by(project_id=project_id).count()
        remaining_motors = Motor.query.filter_by(project_id=project_id).count()

        if remaining_estimates > 0 or remaining_motors > 0:
            flash(f'Warning: Project deleted but {remaining_estimates} estimates and {remaining_motors} motors may remain!', 'warning')
            return redirect(url_for('projects.list_projects'))

        # Build detailed success message
        msg = f'Project "{project_name}" deleted successfully!'
        details = []
        if estimate_count > 0:
            details.append(f'{estimate_count} estimate(s)')
        if assembly_count > 0:
            details.append(f'{assembly_count} assembly/assemblies')
        if assembly_part_count > 0:
            details.append(f'{assembly_part_count} assembly part(s)')
        if component_count > 0:
            details.append(f'{component_count} individual component(s)')
        if motor_count > 0:
            details.append(f'{motor_count} motor(s)/load(s)')

        if details:
            msg += f' Removed: {", ".join(details)}.'

        flash(msg, 'success')
        return redirect(url_for('projects.list_projects'))

    except Exception as e:
        db.session.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f"Error deleting project {project_id}: {error_details}")  # Log to console
        flash(f'Error deleting project: {str(e)}', 'error')
        # Try to redirect to detail page if it still exists, otherwise to list
        try:
            return redirect(url_for('projects.detail_project', project_id=project_id))
        except:
            return redirect(url_for('projects.list_projects'))

@bp.route('/<int:project_id>/toggle-active', methods=['POST'])
def toggle_project_active(project_id):
    """Toggle project active status via AJAX"""
    project = Project.query.get_or_404(project_id)

    try:
        project.is_active = not project.is_active
        project.updated_at = datetime.utcnow()
        db.session.commit()

        return jsonify({
            'success': True,
            'is_active': project.is_active,
            'message': f'Project marked as {"active" if project.is_active else "closed"}'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error updating project: {str(e)}'
        }), 500

@bp.route('/<int:project_id>/copy', methods=['GET', 'POST'])
def copy_project(project_id):
    """Copy an entire project with all its estimates (supporting and optional)"""
    source_project = Project.query.get_or_404(project_id)

    if request.method == 'POST':
        try:
            # Create new project with copied details
            new_project_name = request.form.get('project_name', f"Copy of {source_project.project_name}")

            new_project = Project(
                project_name=new_project_name,
                client_name=source_project.client_name,
                description=source_project.description,
                status='Draft',  # Always start copies as Draft
                revision=request.form.get('revision', ''),
                remarks=source_project.remarks
            )
            db.session.add(new_project)
            db.session.flush()  # Get the new project ID

            # Get all estimates from source project (both supporting and optional)
            source_estimates = Estimate.query.filter_by(project_id=project_id).all()

            # Copy each estimate
            for source_estimate in source_estimates:
                # Create new estimate
                new_estimate_number = f"EST-{datetime.now().strftime('%Y%m%d')}-{str(uuid.uuid4())[:8].upper()}"
                new_estimate = Estimate(
                    project_id=new_project.project_id,
                    estimate_number=new_estimate_number,
                    estimate_name=source_estimate.estimate_name,
                    description=source_estimate.description,
                    is_optional=source_estimate.is_optional,  # Preserve optional status
                    sort_order=source_estimate.sort_order,
                    engineering_rate=source_estimate.engineering_rate,
                    panel_shop_rate=source_estimate.panel_shop_rate,
                    machine_assembly_rate=source_estimate.machine_assembly_rate,
                    engineering_hours=source_estimate.engineering_hours,
                    panel_shop_hours=source_estimate.panel_shop_hours,
                    machine_assembly_hours=source_estimate.machine_assembly_hours
                )
                db.session.add(new_estimate)
                db.session.flush()  # Get the new estimate ID

                # Copy all assemblies and assembly parts
                for assembly in source_estimate.assemblies:
                    new_assembly = Assembly(
                        estimate_id=new_estimate.estimate_id,
                        assembly_name=assembly.assembly_name,
                        description=assembly.description,
                        sort_order=assembly.sort_order,
                        standard_assembly_id=assembly.standard_assembly_id,
                        standard_assembly_version=assembly.standard_assembly_version,
                        quantity=assembly.quantity
                    )
                    db.session.add(new_assembly)
                    db.session.flush()  # Get the new assembly ID

                    # Copy assembly parts
                    for assembly_part in assembly.assembly_parts:
                        new_assembly_part = AssemblyPart(
                            assembly_id=new_assembly.assembly_id,
                            part_id=assembly_part.part_id,
                            quantity=assembly_part.quantity,
                            unit_of_measure=assembly_part.unit_of_measure,
                            sort_order=assembly_part.sort_order,
                            notes=assembly_part.notes
                        )
                        db.session.add(new_assembly_part)

                # Copy individual components (EstimateComponents)
                for individual_component in source_estimate.individual_components:
                    new_individual_component = EstimateComponent(
                        estimate_id=new_estimate.estimate_id,
                        part_id=individual_component.part_id,
                        component_name=individual_component.component_name,
                        part_number=individual_component.part_number,
                        manufacturer=individual_component.manufacturer,
                        description=individual_component.description,
                        unit_price=individual_component.unit_price,
                        quantity=individual_component.quantity,
                        unit_of_measure=individual_component.unit_of_measure,
                        category=individual_component.category,
                        notes=individual_component.notes,
                        sort_order=individual_component.sort_order
                    )
                    db.session.add(new_individual_component)

            # Copy all motors/loads from source project
            source_motors = Motor.query.filter_by(project_id=project_id).order_by(Motor.sort_order).all()
            for source_motor in source_motors:
                new_motor = Motor(
                    project_id=new_project.project_id,
                    load_type=source_motor.load_type,
                    motor_name=source_motor.motor_name,
                    location=source_motor.location,
                    encl_type=source_motor.encl_type,
                    frame=source_motor.frame,
                    additional_notes=source_motor.additional_notes,
                    hp=source_motor.hp,
                    speed_range=source_motor.speed_range,
                    voltage=source_motor.voltage,
                    qty=source_motor.qty,
                    overload_percentage=source_motor.overload_percentage,
                    continuous_load=source_motor.continuous_load,
                    vfd_type_id=source_motor.vfd_type_id,
                    power_rating=source_motor.power_rating,
                    power_unit=source_motor.power_unit,
                    phase_config=source_motor.phase_config,
                    nec_amps_override=source_motor.nec_amps_override,
                    manual_amps=source_motor.manual_amps,
                    vfd_override=source_motor.vfd_override,
                    selected_vfd_part_id=source_motor.selected_vfd_part_id,
                    duty_type=source_motor.duty_type,
                    sort_order=source_motor.sort_order,
                    revision_number='0.0',  # Start fresh with revision 0.0
                    revision_type='major'
                )
                db.session.add(new_motor)

            db.session.commit()

            # Build success message
            success_msg = f'Project "{new_project.project_name}" created successfully with {len(source_estimates)} estimate(s)'
            if source_motors:
                success_msg += f' and {len(source_motors)} motor(s)/load(s)'
            success_msg += '!'

            flash(success_msg, 'success')
            return redirect(url_for('projects.detail_project', project_id=new_project.project_id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error copying project: {str(e)}', 'error')

    # GET request - show the copy form
    return render_template('projects/copy.html', project=source_project)

@bp.route('/<int:project_id>/export')
def export_project(project_id):
    """Export project estimates overview matching the web page format for easy copy-paste"""
    project = Project.query.get_or_404(project_id)

    # Separate standard and optional estimates
    standard_estimates = Estimate.query.filter_by(project_id=project_id, is_optional=False).order_by(Estimate.sort_order, Estimate.created_at.desc()).all()
    optional_estimates = Estimate.query.filter_by(project_id=project_id, is_optional=True).order_by(Estimate.sort_order, Estimate.created_at.desc()).all()
    estimates = standard_estimates + optional_estimates
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{project.project_name} Overview"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    entry_font = Font(bold=True)
    entry_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center')
    right_align = Alignment(horizontal='right')
    
    # Headers exactly matching your image - starting from column B to match hidden column layout
    headers = ['ITEM', 'DESCRIPTION', 'QTY', 'REF', 'ENG / FS', 'WELD', 'MACH', 'ASSY', 'OTHER', 'PURCHASE', 'INVENTORY', 'MAJOR OUTSIDE', 'TOTAL']
    
    # Set column headers starting from column B (2)
    for col, header in enumerate(headers, 2):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    row = 2
    item_number = 24  # Starting from 24 as shown in your image

    # Process standard estimates first
    for estimate in standard_estimates:
        # Add main estimate entry - starting from column B (2)
        ws.cell(row=row, column=2, value=item_number)  # ITEM
        ws.cell(row=row, column=3, value=estimate.estimate_name)  # DESCRIPTION
        ws.cell(row=row, column=4, value=1)  # QTY
        
        # Add labor hours to appropriate columns
        if estimate.engineering_hours > 0:
            ws.cell(row=row, column=6, value=float(estimate.engineering_hours))  # ENG/FS

        # Sum panel shop and machine assembly hours together for ASSY column
        total_assy_hours = float(estimate.panel_shop_hours or 0) + float(estimate.machine_assembly_hours or 0)
        if total_assy_hours > 0:
            ws.cell(row=row, column=9, value=total_assy_hours)  # ASSY
        
        # Add material cost and total
        material_cost = float(estimate.calculated_total)
        total_cost = float(estimate.grand_total)
        
        ws.cell(row=row, column=11, value=material_cost)  # PURCHASE
        ws.cell(row=row, column=14, value=total_cost)  # TOTAL
        
        # Style the main entry row
        ws.cell(row=row, column=3).font = entry_font
        for col in range(2, 15):  # Start from column B (2) to N (14)
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col >= 6:  # Numeric columns
                cell.alignment = right_align
        
        row += 1
        item_number += 1

    # Add separator and process optional estimates if any
    if optional_estimates:
        # Add blank row separator
        row += 1

        # Add "OPTIONAL ESTIMATES" header row
        ws.cell(row=row, column=2, value="OPTIONAL ESTIMATES")
        ws.cell(row=row, column=2).font = Font(bold=True, size=12, color="FF6600")
        ws.cell(row=row, column=2).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for col in range(2, 15):
            ws.cell(row=row, column=col).border = border
        row += 1

        # Process optional estimates
        for estimate in optional_estimates:
            # Add main estimate entry - starting from column B (2)
            ws.cell(row=row, column=2, value=item_number)  # ITEM
            ws.cell(row=row, column=3, value=f"{estimate.estimate_name} (OPTIONAL)")  # DESCRIPTION
            ws.cell(row=row, column=4, value=1)  # QTY

            # Add labor hours to appropriate columns
            if estimate.engineering_hours > 0:
                ws.cell(row=row, column=6, value=float(estimate.engineering_hours))  # ENG/FS

            # Sum panel shop and machine assembly hours together for ASSY column
            total_assy_hours = float(estimate.panel_shop_hours or 0) + float(estimate.machine_assembly_hours or 0)
            if total_assy_hours > 0:
                ws.cell(row=row, column=9, value=total_assy_hours)  # ASSY

            # Add material cost and total
            material_cost = float(estimate.calculated_total)
            total_cost = float(estimate.grand_total)

            ws.cell(row=row, column=11, value=material_cost)  # PURCHASE
            ws.cell(row=row, column=14, value=total_cost)  # TOTAL

            # Style the entry row with highlighting for optional
            ws.cell(row=row, column=3).font = entry_font
            for col in range(2, 15):  # Start from column B (2) to N (14)
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                if col >= 6:  # Numeric columns
                    cell.alignment = right_align

            row += 1
            item_number += 1

    # Set column widths for proper spacing (matching your image format) - starting from column B
    column_widths = [6, 50, 8, 8, 12, 10, 10, 10, 10, 15, 15, 18, 15]
    for col, width in enumerate(column_widths, 2):  # Start from column B (2)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{project.project_name}_Overview.xlsx"'
    
    return response

@bp.route('/<int:project_id>/consolidated-report')
def export_consolidated_report(project_id):
    """Export consolidated BOM report with cost breakdown and pie chart (includes all estimates)"""
    project = Project.query.get_or_404(project_id)
    # Include both standard and optional estimates in BOM report
    standard_estimates = Estimate.query.filter_by(project_id=project_id, is_optional=False).order_by(Estimate.sort_order, Estimate.created_at.desc()).all()
    optional_estimates = Estimate.query.filter_by(project_id=project_id, is_optional=True).order_by(Estimate.sort_order, Estimate.created_at.desc()).all()
    estimates = standard_estimates + optional_estimates
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Define category mapping and colors
    category_mapping = {
        'PANEL': 'Enclosure',
        'ENCLOSURE': 'Enclosure', 
        'MOTOR': 'Power Distribution',
        'CONTACTOR': 'Controls',
        'RELAY': 'Controls',
        'PLC': 'Controls',
        'HMI': 'Controls',
        'SWITCH': 'Controls',
        'FUSE': 'Fuses/Holders',
        'BREAKER': 'Fuses/Holders',
        'WIRE': 'Wiring/Duct',
        'CABLE': 'Wiring/Duct',
        'DUCT': 'Wiring/Duct',
        'TERMINAL': 'Wiring/Duct',
        'SENSOR': 'I/O',
        'INPUT': 'I/O',
        'OUTPUT': 'I/O',
        'ETHERNET': 'Network',
        'NETWORK': 'Network',
        'LABEL': 'Labels',
        'SAFETY': 'Safety',
        'ESTOP': 'Safety'
    }
    
    category_colors = {
        'Enclosure': '1f77b4',
        'Power Distribution': 'ff7f0e', 
        'Controls': '2ca02c',
        'Fuses/Holders': 'd62728',
        'Wiring/Duct': '9467bd',
        'I/O': '8c564b',
        'Network': 'e377c2',
        'Labels': '7f7f7f',
        'Safety': 'bcbd22',
        'Uncategorized': 'ff9896'
    }
    
    # Data cleaning functions
    def clean_text(text):
        if not text:
            return ""
        text = str(text).strip().upper()
        # Fix common OCR artifacts
        text = re.sub(r'PHEONIX', 'PHOENIX', text)
        text = re.sub(r'ENCLOSUER', 'ENCLOSURE', text)
        # Normalize dashes
        text = re.sub(r'[\u2013\u2014\u2015]', '-', text)
        return text
    
    def get_part_category(part):
        """Get category from part database or fallback to keyword mapping"""
        if not part:
            return 'Uncategorized'
        
        # Try to get category from database by querying directly
        if hasattr(part, 'category_id') and part.category_id:
            try:
                category = PartCategory.query.get(part.category_id)
                if category:
                    return category.name
            except:
                pass
        
        # Try the category property
        try:
            if hasattr(part, 'category') and part.category:
                return part.category
        except:
            pass
        
        # Try the relationship directly
        try:
            if hasattr(part, 'part_category') and part.part_category:
                return part.part_category.name
        except:
            pass
        
        # Fallback to keyword mapping if no database category
        text_to_check = f"{getattr(part, 'part_number', '')} {getattr(part, 'description', '')} {getattr(part, 'manufacturer', '')}".upper()
        
        for keyword, category in category_mapping.items():
            if keyword in text_to_check:
                return category
        
        return 'Uncategorized'
    
    def clean_quantity(qty):
        """Clean and normalize quantity"""
        if not qty:
            return 0
        try:
            return max(0, float(qty))  # Treat negatives as credits (0)
        except:
            return 0
    
    # Collect all BOM data
    all_bom_data = []
    category_totals = defaultdict(float)
    missing_data = []
    estimate_totals = {}
    
    for estimate in estimates:
        estimate_total = 0
        estimate_lines = 0
        
        # Process assemblies
        for assembly in estimate.assemblies:
            for assembly_part in assembly.assembly_parts:
                part = assembly_part.part
                
                # Clean data
                qty = clean_quantity(assembly_part.quantity)
                unit_cost = float(part.current_price) if part.current_price else 0
                extended_cost = qty * unit_cost
                
                # Get category from part database
                category = get_part_category(part)
                
                # Track missing data
                if unit_cost == 0:
                    missing_data.append({
                        'estimate_id': estimate.estimate_number,
                        'part_number': part.part_number,
                        'issue': 'Missing unit cost'
                    })
                
                bom_item = {
                    'estimate_id': estimate.estimate_number,
                    'item': len(all_bom_data) + 1,
                    'bb_item': '',
                    'qty': qty,
                    'description': clean_text(part.description),
                    'manufacturer': clean_text(getattr(part, 'manufacturer', '')),
                    'part_number': clean_text(part.part_number),
                    'unit_cost': unit_cost,
                    'extended_cost': extended_cost,
                    'category': category
                }
                
                all_bom_data.append(bom_item)
                category_totals[category] += extended_cost
                estimate_total += extended_cost
                estimate_lines += 1
        
        # Process individual estimate components
        for individual_component in estimate.individual_components:
            # Clean data
            qty = clean_quantity(individual_component.quantity)
            unit_cost = float(individual_component.unit_price) if individual_component.unit_price else 0
            extended_cost = qty * unit_cost
            
            # Get category - use component's category or part's category if linked
            if individual_component.part_id and individual_component.part:
                category = get_part_category(individual_component.part)
                part_number = individual_component.part.part_number
                description = individual_component.part.description
                manufacturer = getattr(individual_component.part, 'manufacturer', '')
            else:
                # Custom component without linked part
                category = individual_component.category if individual_component.category else 'Uncategorized'
                part_number = individual_component.part_number or ''
                description = individual_component.description or individual_component.component_name
                manufacturer = individual_component.manufacturer or ''
            
            # Track missing data
            if unit_cost == 0:
                missing_data.append({
                    'estimate_id': estimate.estimate_number,
                    'part_number': part_number,
                    'issue': 'Missing unit cost'
                })
            
            bom_item = {
                'estimate_id': estimate.estimate_number,
                'item': len(all_bom_data) + 1,
                'bb_item': '',
                'qty': qty,
                'description': clean_text(description),
                'manufacturer': clean_text(manufacturer),
                'part_number': clean_text(part_number),
                'unit_cost': unit_cost,
                'extended_cost': extended_cost,
                'category': category
            }
            
            all_bom_data.append(bom_item)
            category_totals[category] += extended_cost
            estimate_total += extended_cost
            estimate_lines += 1
        
        estimate_totals[estimate.estimate_number] = {
            'total_cost': estimate_total,
            'line_count': estimate_lines
        }
    
    # Create styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center')
    right_align = Alignment(horizontal='right')
    currency_format = '"$"#,##0.00'
    percent_format = '0.0%'
    
    # Sheet 1: Cost Breakdown with Pie Chart
    ws1 = wb.active
    ws1.title = "Cost Breakdown (All Estimates)"
    
    # Title
    ws1.merge_cells('A1:E1')
    ws1['A1'] = f"Cost Breakdown - {project.project_name}"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A1'].alignment = center_align
    
    # Summary table headers
    headers = ['CATEGORY', 'TOTAL COST', 'PERCENT OF TOTAL']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # Calculate totals
    grand_total = sum(category_totals.values())
    
    # Populate summary data
    row = 4
    chart_data = []
    for category in sorted(category_totals.keys(), key=lambda x: category_totals[x], reverse=True):
        total_cost = category_totals[category]
        percent = total_cost / grand_total if grand_total > 0 else 0
        
        ws1.cell(row=row, column=1, value=category).border = border
        ws1.cell(row=row, column=2, value=total_cost).border = border
        ws1.cell(row=row, column=2).number_format = currency_format
        ws1.cell(row=row, column=3, value=percent).border = border
        ws1.cell(row=row, column=3).number_format = percent_format
        
        chart_data.append((category, total_cost))
        row += 1
    
    # Create pie chart
    chart = PieChart()
    chart.title = "Cost Distribution by Category"
    
    # Chart data references
    data = Reference(ws1, min_col=2, min_row=4, max_row=row-1)
    cats = Reference(ws1, min_col=1, min_row=4, max_row=row-1)
    
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    
    # Chart styling
    chart.dataLabels = openpyxl.chart.label.DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showCatName = True
    chart.dataLabels.showLeaderLines = False  # Remove leader lines
    chart.legend.position = 'r'
    
    # Add chart to worksheet
    ws1.add_chart(chart, "G3")
    
    # Missing data section
    if missing_data:
        ws1.cell(row=row+2, column=1, value="Missing Data Issues:").font = Font(bold=True)
        ws1.cell(row=row+3, column=1, value="Estimate ID").font = header_font
        ws1.cell(row=row+3, column=2, value="Part Number").font = header_font
        ws1.cell(row=row+3, column=3, value="Issue").font = header_font
        
        missing_row = row + 4
        for item in missing_data:
            ws1.cell(row=missing_row, column=1, value=item['estimate_id'])
            ws1.cell(row=missing_row, column=2, value=item['part_number'])
            ws1.cell(row=missing_row, column=3, value=item['issue'])
            missing_row += 1
    
    # Sheet 2: All Estimates BOM
    ws2 = wb.create_sheet("All Estimates BOM")
    
    # Headers
    bom_headers = ['ESTIMATE ID', 'ITEM', 'BB ITEM', 'QTY', 'DESCRIPTION', 'MANUFACTURER', 'PART NUMBER', 'UNIT COST', 'EXTENDED COST', 'CATEGORY']
    for col, header in enumerate(bom_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # Add filters to the header row
    ws2.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(bom_headers))}1"
    
    # Freeze the top row
    ws2.freeze_panes = "A2"
    
    # Populate BOM data
    current_estimate = None
    subtotal = 0
    row = 2
    
    # Group data by estimate for easier processing
    estimates_data = {}
    for item in all_bom_data:
        est_id = item['estimate_id']
        if est_id not in estimates_data:
            estimates_data[est_id] = []
        estimates_data[est_id].append(item)
    
    # Process each estimate
    for estimate in estimates:
        if estimate.estimate_number not in estimates_data:
            continue
            
        estimate_items = estimates_data[estimate.estimate_number]
        
        # Add estimate name header row
        ws2.merge_cells(f'A{row}:J{row}')
        estimate_label = f"ESTIMATE: {estimate.estimate_name} ({estimate.estimate_number})"
        if estimate.is_optional:
            estimate_label += " [OPTIONAL]"
        estimate_header = ws2.cell(row=row, column=1, value=estimate_label)
        estimate_header.font = Font(bold=True, size=12, color="FF6600" if estimate.is_optional else "000000")
        estimate_header.fill = PatternFill(start_color="FFF2CC" if estimate.is_optional else "E7E6E6",
                                          end_color="FFF2CC" if estimate.is_optional else "E7E6E6",
                                          fill_type="solid")
        estimate_header.border = border
        row += 1
        
        # Add estimate items
        estimate_subtotal = 0
        for item in estimate_items:
            # Add data row
            ws2.cell(row=row, column=1, value=item['estimate_id'])
            ws2.cell(row=row, column=2, value=item['item'])
            ws2.cell(row=row, column=3, value=item['bb_item'])
            ws2.cell(row=row, column=4, value=item['qty'])
            ws2.cell(row=row, column=5, value=item['description'])
            ws2.cell(row=row, column=6, value=item['manufacturer'])
            ws2.cell(row=row, column=7, value=item['part_number'])
            ws2.cell(row=row, column=8, value=item['unit_cost'])
            ws2.cell(row=row, column=8).number_format = currency_format
            ws2.cell(row=row, column=9, value=item['extended_cost'])
            ws2.cell(row=row, column=9).number_format = currency_format
            ws2.cell(row=row, column=10, value=item['category'])
            
            estimate_subtotal += item['extended_cost']
            row += 1
        
        # Add estimate subtotal
        ws2.cell(row=row, column=8, value="Estimate Subtotal:").font = Font(bold=True)
        ws2.cell(row=row, column=9, value=estimate_subtotal).font = Font(bold=True)
        ws2.cell(row=row, column=9).number_format = currency_format
        row += 1
        
        # Add spacing between estimates
        row += 1
    
    ws2.cell(row=row, column=8, value="GRAND TOTAL:").font = Font(bold=True, size=12)
    ws2.cell(row=row, column=9, value=grand_total).font = Font(bold=True, size=12)
    ws2.cell(row=row, column=9).number_format = currency_format
    
    # Individual estimate sheets
    for estimate in estimates:
        estimate_data = [item for item in all_bom_data if item['estimate_id'] == estimate.estimate_number]
        if not estimate_data:
            continue
            
        # Create sheet for this estimate using estimate name
        # Clean the estimate name for use as sheet name (remove invalid characters)
        clean_name = re.sub(r'[\\/*?:"<>|\[\]]', '_', estimate.estimate_name)
        sheet_name = f"BOM - {clean_name}"[:31]  # Excel sheet name limit
        ws_est = wb.create_sheet(sheet_name)
        
        # Stats section
        unique_parts = len(set(item['part_number'] for item in estimate_data))
        estimate_total = sum(item['extended_cost'] for item in estimate_data)
        
        ws_est.cell(row=1, column=1, value=f"Estimate: {estimate.estimate_name}").font = Font(bold=True, size=14)
        ws_est.cell(row=2, column=1, value=f"Total Lines: {len(estimate_data)}")
        ws_est.cell(row=2, column=3, value=f"Unique Parts: {unique_parts}")
        ws_est.cell(row=3, column=1, value=f"Total Cost: ${estimate_total:,.2f}").font = Font(bold=True)
        
        # BOM headers (without estimate ID)
        est_headers = bom_headers[1:]  # Skip ESTIMATE ID column
        for col, header in enumerate(est_headers, 1):
            cell = ws_est.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        # BOM data
        for row_idx, item in enumerate(estimate_data, 6):
            ws_est.cell(row=row_idx, column=1, value=item['item'])
            ws_est.cell(row=row_idx, column=2, value=item['bb_item'])
            ws_est.cell(row=row_idx, column=3, value=item['qty'])
            ws_est.cell(row=row_idx, column=4, value=item['description'])
            ws_est.cell(row=row_idx, column=5, value=item['manufacturer'])
            ws_est.cell(row=row_idx, column=6, value=item['part_number'])
            ws_est.cell(row=row_idx, column=7, value=item['unit_cost'])
            ws_est.cell(row=row_idx, column=7).number_format = currency_format
            ws_est.cell(row=row_idx, column=8, value=item['extended_cost'])
            ws_est.cell(row=row_idx, column=8).number_format = currency_format
            ws_est.cell(row=row_idx, column=9, value=item['category'])
    
    # Adjust column widths for all sheets
    for ws in wb.worksheets:
        for col_num in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_num)
            
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                # Skip merged cells
                if hasattr(cell, 'value') and cell.value is not None:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
            
            adjusted_width = min(max_length + 2, 50)
            if adjusted_width > 8:  # Only adjust if it's worth it
                ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    today = datetime.now().strftime('%Y%m%d')
    filename = f"Estimate_BOM_Report_{today}.xlsx"
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response